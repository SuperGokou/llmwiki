import base64
import csv
import io
import json
import os
import re
import shutil
import sqlite3
import time
import urllib.error
import urllib.request
from pathlib import Path
from typing import Any, Optional

from flask import Flask, jsonify, render_template, request

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None

try:
    from docx import Document
except ImportError:
    Document = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    from PIL import Image
except ImportError:
    Image = None

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
STATIC_IMAGE_DIR = BASE_DIR / "static" / "images"
LEGACY_DB_PATH = BASE_DIR / "knowledge.db"
DB_PATH = DATA_DIR / "knowledge.db"

DATA_DIR.mkdir(parents=True, exist_ok=True)
STATIC_IMAGE_DIR.mkdir(parents=True, exist_ok=True)
if not DB_PATH.exists() and LEGACY_DB_PATH.exists():
    shutil.copy2(LEGACY_DB_PATH, DB_PATH)


def load_local_env(env_path: Path) -> None:
    """加载本地 .env 到环境变量（不覆盖已存在变量）。"""
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


load_local_env(BASE_DIR / ".env")

OLLAMA_API = os.getenv("OLLAMA_API_URL", "http://127.0.0.1:11434/api/generate")
KNOWLEDGE_MODEL = os.getenv("KNOWLEDGE_MODEL", "gemma4:latest")
OCR_MODEL = os.getenv("OCR_MODEL", "deepseek-ocr:latest")
CHAT_MODEL = os.getenv("CHAT_MODEL", "llama3.2:latest")
REMOTE_API_KEY = os.getenv("API_KEY") or os.getenv("api")
IS_LOCAL_OLLAMA = "127.0.0.1" in OLLAMA_API or "localhost" in OLLAMA_API
OPENAI_API_BASE = os.getenv("OPENAI_API_BASE", "").strip().rstrip("/")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip() or os.getenv("GROQ_API_KEY", "").strip()

# 与 render.yaml 一致：仅声明了 LLM_BACKEND=openai 但未填 Base 时默认 Groq（仍需在 Dashboard 配置密钥）。
if os.getenv("LLM_BACKEND", "").strip().lower() == "openai" and not OPENAI_API_BASE:
    OPENAI_API_BASE = "https://api.groq.com/openai/v1"

OLLAMA_HOST = os.getenv("OLLAMA_HOST", "https://ollama.com").strip().rstrip("/")
OLLAMA_CLOUD_API_KEY = os.getenv("OLLAMA_API_KEY", "").strip()


def is_render_runtime() -> bool:
    """Render 会注入 RENDER=true（参见 Render 官方文档）。"""
    return os.getenv("RENDER", "").strip().lower() in ("true", "1", "yes")


RUNTIME_ENV = "render" if is_render_runtime() else "local"


def use_openai_compatible_backend() -> bool:
    """云端（Render 等）使用 OpenAI 兼容 HTTP API（Groq / OpenAI / OpenRouter 等）。"""
    return os.getenv("LLM_BACKEND", "").strip().lower() == "openai"


def use_ollama_cloud_backend() -> bool:
    """使用 Ollama 官方云（https://ollama.com）等：环境变量 OLLAMA_API_KEY + Client.chat。"""
    if use_openai_compatible_backend():
        return False
    return bool(OLLAMA_CLOUD_API_KEY)


def warn_if_render_misconfigured() -> None:
    if not is_render_runtime():
        return
    if use_openai_compatible_backend():
        if not OPENAI_API_KEY:
            print(
                "RENDER WARNING: LLM_BACKEND=openai 但未设置 OPENAI_API_KEY / GROQ_API_KEY。",
                flush=True,
            )
    elif use_ollama_cloud_backend():
        pass
    elif IS_LOCAL_OLLAMA:
        print(
            "RENDER WARNING: 默认 OLLAMA_API_URL 指向本机，容器内无法访问。"
            "请任选其一：① 设置 OLLAMA_API_KEY（及可选 OLLAMA_HOST=https://ollama.com）使用官方云；"
            "② LLM_BACKEND=openai + OPENAI_API_KEY（Groq/OpenAI）；"
            "③ 将 OLLAMA_API_URL 改为公网自托管 Ollama。",
            flush=True,
        )


warn_if_render_misconfigured()

print(
    f"[LLM-WIKI] runtime={RUNTIME_ENV} RENDER={os.getenv('RENDER')!r} LLM="
    f"{'openai' if use_openai_compatible_backend() else 'ollama_cloud' if use_ollama_cloud_backend() else 'ollama_local'}",
    flush=True,
)

app = Flask(__name__, template_folder="templates")

ALLOWED_EXTENSIONS = {
    ".pdf",
    ".docx",
    ".txt",
    ".md",
    ".csv",
    ".tsv",
    ".xlsx",
    ".png",
    ".jpg",
    ".jpeg",
    ".webp",
    ".bmp",
    ".tiff",
}


def get_db_connection() -> sqlite3.Connection:
    """创建数据库连接，并使用字典风格访问行数据。"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def call_openai_compatible(
    prompt: str,
    model: str,
    images: Optional[list[str]] = None,
    options: Optional[dict[str, Any]] = None,
    keep_alive: Optional[str] = None,
) -> str:
    """OpenAI 兼容 /v1/chat/completions（Groq、OpenAI、OpenRouter 等）。"""
    del keep_alive  # Ollama 专用；兼容接口不使用
    options = options or {}
    temperature = float(options.get("temperature", 0.2))
    max_tokens = int(options.get("num_predict", options.get("max_tokens", 1024)))
    max_tokens = max(16, min(max_tokens, 128000))

    url = f"{OPENAI_API_BASE}/chat/completions"
    if images:
        parts: list[dict[str, Any]] = [{"type": "text", "text": prompt}]
        for b64 in images:
            parts.append(
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}}
            )
        messages = [{"role": "user", "content": parts}]
    else:
        messages = [{"role": "user", "content": prompt}]

    body: dict[str, Any] = {
        "model": model,
        "messages": messages,
        "temperature": temperature,
        "max_tokens": max_tokens,
    }

    req = urllib.request.Request(
        url,
        data=json.dumps(body).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {OPENAI_API_KEY}",
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(req, timeout=180) as response:
            data = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as err:
        detail = err.read().decode("utf-8", errors="ignore")[:800]
        raise ValueError(f"LLM API 异常 (HTTP {err.code}): {detail}") from err
    except urllib.error.URLError as err:
        raise ValueError(f"LLM API 连接失败: {err}") from err

    choices = data.get("choices") if isinstance(data, dict) else None
    if not choices:
        raise ValueError(f"LLM API 返回异常: {str(data)[:400]}")
    msg = choices[0].get("message") or {}
    content = msg.get("content")
    if isinstance(content, list):
        # 少数提供商返回 content 为 parts 列表
        text_parts = []
        for p in content:
            if isinstance(p, dict) and p.get("type") == "text":
                text_parts.append(p.get("text", ""))
        return "".join(text_parts).strip()
    if isinstance(content, str):
        return content.strip()
    return ""


def call_ollama_cloud_chat(
    prompt: str,
    model: str,
    images: Optional[list[str]] = None,
    options: Optional[dict[str, Any]] = None,
    keep_alive: Optional[str] = None,
) -> str:
    """Ollama 官方 HTTP API（云）：Chat Completions，等价于你本地的 Client(host=..., headers=Bearer ...).chat(...)。"""
    try:
        from ollama import Client
    except ImportError as err:
        raise ValueError('请安装依赖: pip install ollama') from err

    host = OLLAMA_HOST or "https://ollama.com"
    client = Client(
        host=host,
        headers={"Authorization": f"Bearer {OLLAMA_CLOUD_API_KEY}"},
    )

    opts: dict[str, Any] = {"temperature": 0.2}
    if options:
        opts.update(options)

    msg: dict[str, Any] = {"role": "user", "content": prompt}
    if images:
        msg["images"] = images

    chat_kw: dict[str, Any] = {
        "model": model,
        "messages": [msg],
        "options": opts,
        "stream": False,
    }
    if keep_alive:
        chat_kw["keep_alive"] = keep_alive

    try:
        resp = client.chat(**chat_kw)
    except Exception as err:
        raise ValueError(f"Ollama Cloud 调用失败: {err}") from err

    if isinstance(resp, dict):
        inner = resp.get("message") or {}
        return str(inner.get("content") or "").strip()
    msg_obj = getattr(resp, "message", None)
    if msg_obj is None:
        return ""
    return str(getattr(msg_obj, "content", "") or "").strip()


def call_ollama(
    prompt: str,
    model: str,
    images: Optional[list[str]] = None,
    options: Optional[dict[str, Any]] = None,
    keep_alive: Optional[str] = None,
) -> str:
    """本地 Ollama /generate；或 LLM_BACKEND=openai；或配置 OLLAMA_API_KEY 走官方云 Client.chat。"""
    if use_openai_compatible_backend():
        if not OPENAI_API_KEY:
            raise ValueError(
                "未检测到 OPENAI_API_KEY（或 GROQ_API_KEY）。请在 Render 控制台为该 Web Service 添加密钥："
                "Dashboard → gokou-llmwiki → Environment → Add Environment Variable，"
                "Key 填 OPENAI_API_KEY，Value 填你在 Groq（https://console.groq.com/keys）"
                "或 OpenAI 等平台生成的 API Key，保存后点 Manual Deploy 重新部署。"
            )
        if not OPENAI_API_BASE:
            raise ValueError(
                "OPENAI_API_BASE 为空。请在 Environment 设置，例如 https://api.groq.com/openai/v1"
            )
        return call_openai_compatible(
            prompt=prompt,
            model=model,
            images=images,
            options=options,
            keep_alive=keep_alive,
        )

    if use_ollama_cloud_backend():
        return call_ollama_cloud_chat(
            prompt=prompt,
            model=model,
            images=images,
            options=options,
            keep_alive=keep_alive,
        )

    payload = {
        "model": model,
        "prompt": prompt,
        "stream": False,
        "options": {"temperature": 0.2},
    }
    if options:
        payload["options"].update(options)
    if images:
        payload["images"] = images
    if keep_alive:
        payload["keep_alive"] = keep_alive

    max_attempts = 2
    for attempt in range(max_attempts):
        req = urllib.request.Request(
            OLLAMA_API,
            data=json.dumps(payload).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        if REMOTE_API_KEY and not IS_LOCAL_OLLAMA:
            # 兼容未来切换远端 API 网关的鉴权头。
            req.add_header("Authorization", f"Bearer {REMOTE_API_KEY}")
            req.add_header("x-api-key", REMOTE_API_KEY)

        try:
            with urllib.request.urlopen(req, timeout=180) as response:
                data = json.loads(response.read().decode("utf-8"))
                return data.get("response", "")
        except urllib.error.HTTPError as err:
            error_body = ""
            try:
                error_body = err.read().decode("utf-8", errors="ignore")
                parsed = json.loads(error_body) if error_body else {}
                error_body = parsed.get("error", "") if isinstance(parsed, dict) else error_body
            except Exception:  # noqa: BLE001
                pass

            # Ollama 偶发 5xx 时做一次快速重试，提升稳定性。
            if err.code >= 500 and attempt < max_attempts - 1:
                time.sleep(1.0)
                continue
            detail = (error_body or str(err)).strip()[:300]
            raise ValueError(f"Ollama 服务异常 (HTTP {err.code}): {detail}") from err
        except urllib.error.URLError as err:
            if attempt < max_attempts - 1:
                time.sleep(1.0)
                continue
            raise ValueError(f"Ollama 连接失败，请确认服务已启动: {err}") from err
        except Exception as err:  # noqa: BLE001
            raise ValueError(f"Ollama 调用失败: {err}") from err

    raise ValueError("Ollama 调用失败：超过最大重试次数")


def extract_json_block(raw_text: str) -> str:
    """从 LLM 文本中提取 JSON 代码块。"""
    fenced = re.search(r"```json\s*(.*?)\s*```", raw_text, re.DOTALL | re.IGNORECASE)
    if fenced:
        return fenced.group(1).strip()
    array_block = re.search(r"(\[\s*{.*}\s*\])", raw_text, re.DOTALL)
    if array_block:
        return array_block.group(1).strip()
    return raw_text.strip()


def extract_text_with_deepseek_ocr(image_bytes: bytes) -> str:
    """使用 deepseek-ocr 提取图片文字。"""
    image_b64 = base64.b64encode(image_bytes).decode("utf-8")
    prompt = (
        "请提取图片中的全部可读文字，按自然阅读顺序输出纯文本。"
        "不要解释，不要添加额外内容。"
    )
    text = call_ollama(prompt=prompt, model=OCR_MODEL, images=[image_b64]).strip()
    return text


def extract_text_from_scanned_pdf(file_bytes: bytes, max_pages: int = 10) -> str:
    """将扫描版 PDF 渲染为图片后走 deepseek-ocr。"""
    if fitz is None:
        raise ValueError("缺少 pymupdf，无法处理扫描版 PDF。")
    document = fitz.open(stream=file_bytes, filetype="pdf")
    page_count = min(document.page_count, max_pages)
    chunks: list[str] = []
    for page_index in range(page_count):
        page = document.load_page(page_index)
        pix = page.get_pixmap(dpi=220)
        page_png = pix.tobytes("png")
        page_text = extract_text_with_deepseek_ocr(page_png).strip()
        if page_text:
            chunks.append(page_text)
    document.close()
    return "\n".join(chunks)


def extract_text_from_file(file_name: str, file_bytes: bytes) -> str:
    """按文件扩展名抽取文本内容，OCR 优先使用 deepseek-ocr。"""
    suffix = Path(file_name).suffix.lower()

    if suffix in {".txt", ".md"}:
        return file_bytes.decode("utf-8", errors="ignore")

    if suffix in {".csv", ".tsv"}:
        delimiter = "," if suffix == ".csv" else "\t"
        text = file_bytes.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = [" | ".join(col.strip() for col in row if col) for row in reader]
        return "\n".join(row for row in rows if row.strip())

    if suffix == ".docx":
        if Document is None:
            raise ValueError("缺少 python-docx 依赖，请安装: pip install python-docx")
        document = Document(io.BytesIO(file_bytes))
        return "\n".join(p.text for p in document.paragraphs)

    if suffix == ".xlsx":
        if load_workbook is None:
            raise ValueError("缺少 openpyxl 依赖，请安装: pip install openpyxl")
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"[工作表] {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                line = " | ".join(str(cell).strip() for cell in row if cell is not None)
                if line:
                    lines.append(line)
        return "\n".join(lines)

    if suffix == ".pdf":
        if PdfReader is None:
            raise ValueError("缺少 pypdf 依赖，请安装: pip install pypdf")
        reader = PdfReader(io.BytesIO(file_bytes))
        extracted = "\n".join((page.extract_text() or "") for page in reader.pages).strip()
        if extracted:
            return extracted
        scanned = extract_text_from_scanned_pdf(file_bytes).strip()
        if scanned:
            return scanned
        raise ValueError(
            "PDF 未提取到文本内容。可能是扫描件，请确认 deepseek-ocr:latest 模型可用。"
        )

    if suffix in {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tiff"}:
        ocr_text = extract_text_with_deepseek_ocr(file_bytes).strip()
        if ocr_text:
            return ocr_text
        raise ValueError("图片 OCR 失败，请检查 deepseek-ocr:latest 是否可调用。")

    raise ValueError(f"不支持的文件类型: {suffix}")


def extract_knowledge_points_with_llm(
    source_title: str, source_text: str, max_points: int = 8
) -> list[dict[str, str]]:
    """用 gemma4 抽取结构化知识点。"""
    compact_text = " ".join(source_text.split())
    prompt = f"""
你是知识工程助手。请从以下文本中抽取 {max_points} 个以内“可独立成节点”的知识点。
返回严格 JSON 数组，不要输出任何额外文字。
每个元素格式:
{{
  "name": "知识点名称",
  "summary": "60-120字摘要"
}}

标题: {source_title}
正文:
{compact_text[:12000]}
""".strip()

    raw = call_ollama(prompt=prompt, model=KNOWLEDGE_MODEL)
    json_text = extract_json_block(raw)
    parsed = json.loads(json_text)
    if not isinstance(parsed, list):
        raise ValueError("知识点抽取结果不是数组。")

    points: list[dict[str, str]] = []
    for item in parsed:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name", "")).strip()
        summary = str(item.get("summary", "")).strip()
        if name and summary:
            points.append({"name": name[:120], "summary": summary[:600]})
    if points:
        return points

    # LLM 返回异常时，保底生成一个知识点。
    fallback_summary = (compact_text[:220] or source_title).strip()
    return [{"name": source_title[:120], "summary": fallback_summary}]


def generate_predicted_questions(point_name: str, point_summary: str) -> list[tuple[str, str]]:
    """基于知识点生成逆向问题（当前保持轻量模板，后续可再 LLM 化）。"""
    hint = point_summary[:80] if point_summary else point_name
    return [
        (f"{point_name} 在真实客户场景中解决的首要问题是什么？", "价值认知"),
        (f"围绕“{point_name}”，客户最常提出的实施难点有哪些？", "实施策略"),
        (f"如果“{point_name}”效果不达预期，第一步应如何排查？", "排障优化"),
        (f"基于“{hint}”，客户下一步最可能追问什么？", "追问预测"),
    ]


def _build_query_terms(user_message: str) -> set[str]:
    """构建更稳健的查询关键词（含中英文词和中文双字词）。"""
    normalized = (user_message or "").lower()
    terms: set[str] = set()

    # 英文/数字词块。
    terms.update(re.findall(r"[a-z0-9_]{2,}", normalized))

    # 连续中文段。
    chinese_chunks = re.findall(r"[\u4e00-\u9fa5]+", normalized)
    for chunk in chinese_chunks:
        if len(chunk) <= 1:
            continue
        # 整段保留一次（便于完整匹配）。
        terms.add(chunk)
        # 生成双字词，避免“整句一个 token”导致命中率过低。
        for i in range(len(chunk) - 1):
            terms.add(chunk[i : i + 2])

    # 去掉过短噪声。
    return {t for t in terms if len(t.strip()) >= 2}


def retrieve_relevant_rows(
    rows: list[sqlite3.Row], user_message: str, top_k: int = 5
) -> tuple[list[sqlite3.Row], int]:
    """按关键词重合度返回相关条目与最高分。"""
    query_terms = _build_query_terms(user_message)

    def score_row(row: sqlite3.Row) -> int:
        title = (row["title"] or "").lower()
        content = (row["content"] or "").lower()
        haystack = f"{title} {content[:320]}"
        if not query_terms:
            return 1
        return sum(1 for term in query_terms if term and term in haystack)

    scored = [(score_row(row), row) for row in rows]
    scored.sort(key=lambda x: x[0], reverse=True)

    top_score = scored[0][0] if scored else 0
    # 当相关度很低时，自动扩大上下文，避免“答不出来”。
    effective_top_k = top_k if top_score > 0 else min(8, len(rows))
    selected = [row for _, row in scored[:effective_top_k]]
    return selected, top_score


def select_relevant_context(rows: list[sqlite3.Row], user_message: str, top_k: int = 5) -> str:
    """按关键词重合度挑选更相关的知识片段，减少提示词体积。"""
    selected, _ = retrieve_relevant_rows(rows, user_message=user_message, top_k=top_k)
    chunks: list[str] = []
    for row in selected:
        title = row["title"]
        content = (row["content"] or "")[:260]
        chunks.append(f"- {title}: {content}")
    return "\n".join(chunks)


def looks_like_low_confidence_reply(text: str) -> bool:
    """判断回复是否是“无法回答/信息不足”类型，用于触发慢模型兜底。"""
    lowered = (text or "").lower()
    markers = [
        "无法提供",
        "无法回答",
        "信息不足",
        "没有相关信息",
        "知识库中没有",
        "不确定",
        "不清楚",
    ]
    return any(marker in lowered for marker in markers)


@app.route("/")
def index():
    """渲染知识图谱页面。"""
    return render_template("graph.html", runtime_env=RUNTIME_ENV)


@app.route("/api/runtime")
def api_runtime():
    """前端或运维探测当前进程运行在本地还是 Render。"""
    if use_openai_compatible_backend():
        llm_mode = "openai_compatible"
    elif use_ollama_cloud_backend():
        llm_mode = "ollama_cloud"
    else:
        llm_mode = "ollama_local"
    return jsonify(
        {
            "environment": RUNTIME_ENV,
            "is_render": is_render_runtime(),
            "llm_backend": llm_mode,
        }
    )


@app.route("/api/graph")
def get_graph():
    """返回图谱所需的节点和边数据。"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # wiki_articles 在当前版本中承载“知识点节点”数据。
        cursor.execute("SELECT id, title FROM wiki_articles ORDER BY id")
        article_rows = cursor.fetchall()

        cursor.execute(
            "SELECT id, article_id, question, question_type FROM predicted_questions ORDER BY id"
        )
        question_rows = cursor.fetchall()

        # 文章节点：id 使用 article-{id}，类型固定为 article。
        nodes: list[dict[str, Any]] = [
            {
                "id": f"article-{row['id']}",
                "raw_id": row["id"],
                "type": "article",
                "label": row["title"],
            }
            for row in article_rows
        ]

        # 问题节点：id 使用 question-{id}，并携带问题类型便于前端扩展展示。
        question_nodes = [
            {
                "id": f"question-{row['id']}",
                "raw_id": row["id"],
                "type": "question",
                "label": row["question"],
                "question_type": row["question_type"],
                "article_id": row["article_id"],
            }
            for row in question_rows
        ]
        nodes.extend(question_nodes)

        # 边：从文章节点指向问题节点，来源是 predicted_questions.article_id。
        links = [
            {
                "source": f"article-{row['article_id']}",
                "target": f"question-{row['id']}",
            }
            for row in question_rows
        ]

        return jsonify({"nodes": nodes, "links": links})
    except sqlite3.Error as db_error:
        return jsonify({"error": f"数据库读取失败: {db_error}"}), 500
    finally:
        if "conn" in locals():
            conn.close()


@app.route("/api/uploaded-files")
def get_uploaded_files():
    """返回已上传文件列表（按源文件聚合）。"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT
                source,
                COUNT(*) AS point_count,
                MAX(id) AS latest_article_id
            FROM wiki_articles
            WHERE source LIKE 'upload:%'
            GROUP BY source
            ORDER BY latest_article_id DESC
            """
        )
        rows = cursor.fetchall()
        files = [
            {
                "source": row["source"],
                "file_name": row["source"].replace("upload:", "", 1),
                "point_count": row["point_count"],
                "latest_article_id": row["latest_article_id"],
            }
            for row in rows
        ]
        return jsonify({"files": files})
    except sqlite3.Error as db_error:
        return jsonify({"error": f"数据库读取失败: {db_error}"}), 500
    finally:
        if "conn" in locals():
            conn.close()


@app.route("/api/article/<int:article_id>")
def get_article(article_id: int):
    """按文章 ID 返回详细内容。"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, title, content, source FROM wiki_articles WHERE id = ?",
            (article_id,),
        )
        row = cursor.fetchone()

        if row is None:
            return jsonify({"error": "文章不存在"}), 404

        return jsonify(
            {
                "id": row["id"],
                "title": row["title"],
                "content": row["content"],
                "source": row["source"],
            }
        )
    except sqlite3.Error as db_error:
        return jsonify({"error": f"数据库读取失败: {db_error}"}), 500
    finally:
        if "conn" in locals():
            conn.close()


@app.route("/api/article/<int:article_id>", methods=["DELETE"])
def delete_article(article_id: int):
    """删除知识点及其关联问题。"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT id, title FROM wiki_articles WHERE id = ?", (article_id,))
        article_row = cursor.fetchone()
        if article_row is None:
            return jsonify({"error": "文章不存在"}), 404

        cursor.execute("DELETE FROM predicted_questions WHERE article_id = ?", (article_id,))
        deleted_question_count = cursor.rowcount
        cursor.execute("DELETE FROM wiki_articles WHERE id = ?", (article_id,))
        deleted_article_count = cursor.rowcount
        conn.commit()

        return jsonify(
            {
                "message": "删除成功",
                "deleted_article_id": article_id,
                "deleted_article_count": deleted_article_count,
                "deleted_question_count": deleted_question_count,
            }
        )
    except sqlite3.Error as db_error:
        return jsonify({"error": f"数据库删除失败: {db_error}"}), 500
    finally:
        if "conn" in locals():
            conn.close()


@app.route("/api/chat", methods=["POST"])
def chat_with_knowledge_base():
    """右下角聊天机器人接口：结合知识库内容回答问题。"""
    payload = request.get_json(silent=True) or {}
    user_message = str(payload.get("message", "")).strip()
    if not user_message:
        return jsonify({"error": "消息不能为空"}), 400

    try:
        started_at = time.perf_counter()
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT title, content
            FROM wiki_articles
            ORDER BY id DESC
            LIMIT 12
            """
        )
        rows = cursor.fetchall()
        if not rows:
            return jsonify({"reply": "当前知识库为空，请先上传文件再提问。"})

        selected_rows, top_score = retrieve_relevant_rows(rows, user_message=user_message, top_k=5)
        context = "\n".join(
            f"- {row['title']}: {(row['content'] or '')[:260]}"
            for row in selected_rows
        )

        prompt = f"""
你是“LLM-WIKI 知识库助手”。请基于给定知识点回答用户问题。
要求：
1) 优先引用知识库内容，不要编造。
2) 用中文，简洁清晰。
3) 如果知识库信息不足，请明确说明并建议补充哪些资料。

知识库片段：
{context}

用户问题：
{user_message}
""".strip()

        reply = call_ollama(
            prompt=prompt,
            model=CHAT_MODEL,
            options={
                "temperature": 0.15,
                "num_predict": 520,
                "num_ctx": 2048,
            },
            keep_alive="30m",
        ).strip()

        # 快模型答不出来时，自动用主模型 + 更大上下文兜底一次。
        if looks_like_low_confidence_reply(reply):
            fallback_context = select_relevant_context(rows, user_message=user_message, top_k=8)
            fallback_prompt = f"""
你是“LLM-WIKI 知识库助手”。请仅根据给定知识库片段回答用户问题。
要求：
1) 先给出结论，再给出依据（引用片段中的关键点）。
2) 不要编造知识库外事实。
3) 若知识库确实缺失，明确指出缺少哪类信息。

知识库片段：
{fallback_context}

用户问题：
{user_message}
""".strip()
            fallback_reply = call_ollama(
                prompt=fallback_prompt,
                model=KNOWLEDGE_MODEL,
                options={
                    "temperature": 0.1,
                    "num_predict": 700,
                    "num_ctx": 4096,
                },
                keep_alive="30m",
            ).strip()
            if fallback_reply:
                reply = fallback_reply

        # 若回答看起来被截断（结尾没有完整句号/问号/叹号），补一段续写。
        if reply and not re.search(r"[。！？.!?]\s*$", reply):
            continue_prompt = f"""
请继续上一个回答，只输出后续内容，不要重复前文。
上一个回答：
{reply}
""".strip()
            continuation = call_ollama(
                prompt=continue_prompt,
                model=CHAT_MODEL,
                options={
                    "temperature": 0.1,
                    "num_predict": 220,
                    "num_ctx": 2048,
                },
                keep_alive="30m",
            ).strip()
            if continuation:
                reply = f"{reply}\n{continuation}".strip()

        # 仍低置信时，若检索命中则给出基于知识库片段的结构化保底回答。
        if looks_like_low_confidence_reply(reply) and top_score > 0 and selected_rows:
            top_titles = [row["title"] for row in selected_rows[:3]]
            top_snippets = [
                (row["content"] or "").strip().replace("\n", " ")[:120]
                for row in selected_rows[:2]
            ]
            evidence = "\n".join(f"- {snippet}" for snippet in top_snippets if snippet)
            title_line = "、".join(top_titles)
            reply = (
                f"根据当前知识库，与你问题最相关的条目是：{title_line}。\n"
                f"可提炼的信息如下：\n{evidence}\n"
                "如果你需要“政策条目（资金补贴、人才政策、园区支持、时间节点）”的完整清单，"
                "可以继续让我按该主题输出结构化要点。"
            )

        latency_ms = int((time.perf_counter() - started_at) * 1000)
        return jsonify(
            {
                "reply": reply or "暂时无法生成回答，请稍后重试。",
                "model": CHAT_MODEL,
                "latency_ms": latency_ms,
            }
        )
    except Exception as err:  # noqa: BLE001
        return jsonify({"error": f"聊天接口调用失败: {err}"}), 500
    finally:
        if "conn" in locals():
            conn.close()


@app.route("/api/upload-knowledge", methods=["POST"])
def upload_knowledge_files():
    """上传知识库文件并自动入库，随后生成预测问题。"""
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "请至少上传一个文件"}), 400

    imported: list[dict[str, Any]] = []
    failed: list[dict[str, str]] = []

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        for uploaded in files:
            file_name = uploaded.filename or ""
            suffix = Path(file_name).suffix.lower()
            if not file_name:
                failed.append({"file": "unknown", "reason": "文件名为空"})
                continue
            if suffix not in ALLOWED_EXTENSIONS:
                failed.append(
                    {
                        "file": file_name,
                        "reason": "仅支持 PDF、Word、TXT、Markdown、CSV/TSV、Excel、图片(png/jpg/jpeg/webp/bmp/tiff)",
                    }
                )
                continue

            try:
                file_bytes = uploaded.read()
                extracted_text = extract_text_from_file(file_name, file_bytes).strip()
                if not extracted_text:
                    raise ValueError("文件未解析到有效文本内容")

                source_title = Path(file_name).stem
                knowledge_points = extract_knowledge_points_with_llm(
                    source_title=source_title,
                    source_text=extracted_text,
                    max_points=8,
                )

                generated_questions_count = 0
                for point in knowledge_points:
                    point_name = point["name"]
                    point_summary = point["summary"]
                    source = f"upload:{file_name}"
                    cursor.execute(
                        "INSERT INTO wiki_articles (title, content, source) VALUES (?, ?, ?)",
                        (point_name, point_summary, source),
                    )
                    point_article_id = cursor.lastrowid

                    generated_questions = generate_predicted_questions(point_name, point_summary)
                    generated_questions_count += len(generated_questions)
                    cursor.executemany(
                        """
                        INSERT INTO predicted_questions (article_id, question, question_type)
                        VALUES (?, ?, ?)
                        """,
                        [(point_article_id, q, q_type) for q, q_type in generated_questions],
                    )

                imported.append(
                    {
                        "file": file_name,
                        "knowledge_points": len(knowledge_points),
                        "generated_questions": generated_questions_count,
                        "extract_model": KNOWLEDGE_MODEL,
                        "ocr_model": OCR_MODEL,
                    }
                )
            except ValueError as err:
                failed.append({"file": file_name, "reason": str(err)})
            except Exception as err:  # noqa: BLE001
                failed.append({"file": file_name, "reason": f"解析失败: {err}"})

        conn.commit()
        return jsonify(
            {
                "imported_count": len(imported),
                "failed_count": len(failed),
                "imported": imported,
                "failed": failed,
            }
        )
    except sqlite3.Error as db_error:
        return jsonify({"error": f"数据库写入失败: {db_error}"}), 500
    finally:
        if "conn" in locals():
            conn.close()


if __name__ == "__main__":
    # 通过环境变量控制 debug，避免默认暴露调试器。
    debug_mode = os.getenv("FLASK_DEBUG", "0") == "1"
    port = int(os.getenv("PORT", "5000"))
    host = os.getenv("BIND_HOST", "127.0.0.1")
    app.run(host=host, port=port, debug=debug_mode)
